# -*- coding: utf-8 -*-
"""Data grouping, averaging, and Excel output helpers for HB Automation."""

from __future__ import annotations

from pathlib import Path
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter


def preprocessing(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize legacy IPEmotion CSV and converted IAD CSV data."""
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]

    unit_row = None
    for index in range(min(len(df), 80)):
        row_text = " ".join(map(str, df.iloc[index].tolist()))
        if "(s)" in row_text:
            unit_row = index
            break

    if unit_row is not None and unit_row >= 1:
        header_row = max(unit_row - 2, 0)
        df.columns = df.iloc[header_row].astype(str)
        df = df.iloc[header_row + 1 :].reset_index(drop=True)

        unit_index = unit_row - header_row - 1
        if 0 <= unit_index < len(df):
            df = df.drop(index=unit_index)
        df.reset_index(drop=True, inplace=True)

    for column in df.columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    df = df.dropna(axis=0, how="all")
    return df.dropna(axis=1, how="all")


def determine_time_step(time_series: pd.Series) -> float:
    return float(np.nanmean(np.diff(time_series)))


def first_type(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    grouped_ones: dict[str, pd.Series] = {}
    grouped_tens: dict[str, pd.Series] = {}
    grouped_hundreds: dict[str, pd.Series] = {}

    time_columns = [column for column in df.columns if str(column).endswith(".X")]
    time_step_map = {}
    for column in time_columns:
        time_step = determine_time_step(df[column])
        time_step_map[column] = time_step

        if 0.9 < time_step < 1.1:
            grouped_ones["Time_1Hz"] = df[column]
        elif 0.09 < time_step < 0.11:
            grouped_tens["Time_10Hz"] = df[column]
        elif 0.009 < time_step < 0.011:
            grouped_hundreds["Time_100Hz"] = df[column]

    for column in df.columns:
        if column in time_step_map:
            continue
        for _, time_step in time_step_map.items():
            if 0.9 < time_step < 1.1:
                grouped_ones[column] = df[column]
            elif 0.09 < time_step < 0.11:
                grouped_tens[column] = df[column]
            elif 0.009 < time_step < 0.011:
                grouped_hundreds[column] = df[column]

    return pd.DataFrame(grouped_ones), pd.DataFrame(grouped_tens), pd.DataFrame(grouped_hundreds)


def second_type(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    grouped_ones: dict[str, pd.Series] = {}
    grouped_tens: dict[str, pd.Series] = {}
    grouped_hundreds: dict[str, pd.Series] = {}

    columns = list(df.columns)
    index = 0
    while index < len(columns):
        column = str(columns[index]).strip()
        match = re.match(r"Time[_ ]?(\d+)Hz", column)
        if not match:
            index += 1
            continue

        hz = int(match.group(1))
        if hz == 1:
            target = grouped_ones
        elif hz == 10:
            target = grouped_tens
        elif hz == 100:
            target = grouped_hundreds
        else:
            index += 1
            continue

        target[column] = df[columns[index]]
        next_index = index + 1
        while next_index < len(columns):
            next_column = str(columns[next_index]).strip()
            if re.match(r"Time[_ ]?\d+Hz", next_column):
                break
            target[next_column] = df[columns[next_index]]
            next_index += 1
        index = next_index

    return pd.DataFrame(grouped_ones), pd.DataFrame(grouped_tens), pd.DataFrame(grouped_hundreds)


def average_channel(groups: dict[int, pd.DataFrame], channel_name: str, start_s: float, end_s: float) -> float:
    if start_s > end_s:
        start_s, end_s = end_s, start_s

    for rate, df in groups.items():
        if df.empty or channel_name not in df.columns:
            continue
        time_column = df.columns[0]
        filtered = df.loc[(df[time_column] >= start_s) & (df[time_column] <= end_s), channel_name]
        numeric = pd.to_numeric(filtered, errors="coerce").dropna()
        if numeric.empty:
            raise ValueError(f"Channel has no numeric samples in averaging range: {channel_name}")
        return round(float(numeric.mean()), 1)

    raise KeyError(f"Channel not found: {channel_name}")


def export_average_excel(
    output_path: str | Path,
    header_df: pd.DataFrame,
    sub_df: pd.DataFrame,
    main_df: pd.DataFrame,
    start_s: float,
    end_s: float,
    source_file: str | Path | None = None,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    temp_path = output_path.with_suffix(".tmp.xlsx")
    top_rows = len(header_df) + len(sub_df)
    df = pd.concat([header_df, sub_df, main_df], axis=0)
    df.to_excel(temp_path, index=False, header=False)

    wb = load_workbook(temp_path)
    ws = wb.active
    comment_lines = [
        "Averaging time:",
        f"{min(start_s, end_s)} ~ {max(start_s, end_s)}",
    ]
    if source_file:
        comment_lines.extend(["Source file:", Path(source_file).name])
    ws["B1"].comment = Comment(text="\n".join(comment_lines), author="HB Automation")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    column_widths: dict[int, float] = {}
    for col in ws.columns:
        max_length = 0
        column_index = col[0].column
        column = get_column_letter(column_index)
        for cell in col:
            max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        column_widths[column_index] = max_length + 2
        ws.column_dimensions[column].width = column_widths[column_index]

    value_column_width = max(column_widths.get(2, 0), column_widths.get(3, 0), 10)
    for row in range(1, top_rows + 1):
        if ws.cell(row, 3).value is not None and ws.cell(row, 2).value in (None, ""):
            ws.cell(row, 2).value = ws.cell(row, 3).value
        ws.cell(row, 3).value = None
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 2).border = thin_border

    ws.column_dimensions["B"].width = value_column_width
    ws.column_dimensions["C"].width = value_column_width

    wb.save(output_path)
    wb.close()
    temp_path.unlink(missing_ok=True)
    return output_path
